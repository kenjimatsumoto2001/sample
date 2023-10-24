# jsonディレクトリ以下にあるmashup_data.json, api_data.json　を用いて，
# 各サービスのカテゴリ情報や，マッシュアップしたサービスなどを取得できる
import json
import collections

def load_json(file_pass):
    fw = open(file_pass, "r")
    json_data = json.load(fw)
    return json_data

def write_json(data, file_pass, indent=4):
    with open(file_pass, 'w') as f:
        json.dump(data, f, indent=indent)

class Get_Api: ## apiデータの操作
    def __init__(self, api_pass):
        self.api_json = load_json(api_pass)

    ### api_nameの代表カテゴリを返す，代表カテゴリはカテゴリリストの先頭とする
    def get_category(self, api_name: str = "google_maps") -> str: 
        try:
            if len(self.api_json[api_name][0]) == 0:
                return "No_category"
            else:
                return self.api_json[api_name][0][0]
        except KeyError:
            return "No_category"

    ### api_nameのカテゴリ一覧を返す
    def get_category_list(self, api_name: str = "google_maps") -> list:
        try:
            return self.api_json[api_name][0]
        except KeyError:
            return []
    ### api_nameの説明文を返す
    def get_describe(self, api_name: str = "google_maps") -> str:
        try:
            return self.api_json[api_name][1]
        except KeyError:
            return "no_describe"
    ###apiの提供者(provider)を返す
    def get_privider(self, api_name: str = "google_maps") -> str:
        try:
            return self.api_json[api_name][2]
        except KeyError:
            return "no_describe"

    ### {api名：カテゴリ１つ}を返す
    def get_all_api_category(self) -> dict:
        return {key:val[0][0] for key, val in self.api_json.items() if len(val[0]) > 0}

    ### {api名：全カテゴリ}を返す
    def get_all_api_all_category(self) -> dict:
        return {key:val[0] for key, val in self.api_json.items() if len(val[0]) > 0}

class Get_Mashup: ## マッシュアップデータの操作
    def __init__(self, mashup_pass):
        self.mashup_json = load_json(mashup_pass)

    ### mashup_nameの代表カテゴリを返す
    def get_category(self, mashup_name: str = "digireality") -> str:
        return self.mashup_json[mashup_name][0][0]

    ### mashup_nameのカテゴリ一覧を返す
    def get_category_list(self, mashup_name: str = "digireality") -> list:
        return self.mashup_json[mashup_name][0]

    ### mashup_nameの説明文を返す
    def get_describe(self, mashup_name: str = "digireality") -> str:
        return self.mashup_json[mashup_name][1]

    ### mashup_nameがマッシュアップしたAPI一覧を返す
    def get_mashup_apis(self, mashup_name: str = "digireality") -> list:
        return self.mashup_json[mashup_name][2]

    ### {マッシュアップ名：APIリスト}を返す，APIを２つ以上マッシュアップするものをマッシュアップサービスとする
    def get_all_mashup_apis(self) -> dict:
        return {k:v[2] for k, v in self.mashup_json.items() if len(v[2]) >= 2}

if __name__=='__main__':
    import pprint
    api_pass = "./json/api_data_new.json"
    mashup_pass = "./json/mashup_data_new.json"
    get_api = Get_Api(api_pass)
    get_mashup = Get_Mashup(mashup_pass)
    mashup_dict = get_mashup.get_all_mashup_apis()
    print(get_api.get_privider("mind42"))
    #print(mashup_dict)
    exit()
    edge_dic = {}
    for i in mashup_dict.values():
        for g in i:
            if g not in edge_dic:
                edge_dic[g] = []
                for j in i:
                    if j == g:
                        pass
                    else:
                        edge_dic[g].append(j)
            else:
                for h in i:
                    if h == g:
                        pass
                    else:
                        edge_dic[g].append(h)
    
    tes = edge_dic["trynt"]
    #print(tes)
    #以下、重複を消す
    #tes = [s for s in tes if s != 'yahoo_search']
    #print(tes)
    #以下、要素のカウント
    num_c = tes.count('yahoo_search')
    #print(num_c)
    c = collections.Counter(tes)
    print(c)
    
    print(list(c.keys())[0])
    exit()
    
    mashup_dict = get_mashup.get_all_mashup_apis()
    used_apis = list()
    for row in mashup_dict.values():
        used_apis.extend(row)
    used_apis = list(set(used_apis))
    used_api_dict = {row:get_api.get_category_list(row) for row in used_apis}
    for row in used_apis:
        if len(get_api.get_category_list(row)) > 5:
            print(row, get_api.get_category_list(row))
    exit()
    
    #縦：サービス数，横：連携数　の度数分布を作る
    import openpyxl
    excel_name = "./excel/api_category_degree_graph.xlsx"
    #wb = openpyxl.load_workbook(excel_name)
    wb = openpyxl.Workbook()
    ws = wb.active
    #ws = wb['Sheet1']
    for i, row in enumerate(used_api_dict.items()):
        ws.cell(i+1, 1, value=row[0])
        ws.cell(i+1, 2, value=len(row[1]))
    wb.save(excel_name)
    
    


    

